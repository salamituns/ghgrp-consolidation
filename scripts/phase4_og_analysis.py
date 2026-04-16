"""
Phase 4a — Oil & Gas Sector (GHGRP Subpart W)
Analysis and rendering of O&G facility ownership concentration.
"""
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import geopandas as gpd
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
import json
import openpyxl
from pyxlsb import open_workbook
from shapely.geometry import shape, Point

GEOWORKS = '/sessions/wonderful-amazing-volta/mnt/CoWorker/Geoworks'
RAW_EPA  = f'{GEOWORKS}/data/raw/epa_ghgrp'
RAW_CB   = f'{GEOWORKS}/data/raw/census_boundaries'
PROC     = f'{GEOWORKS}/data/processed'
VIZ      = f'{GEOWORKS}/viz'

# ---------- 1. Identify Subpart W (Oil & Gas) facilities ----------
PATH = f'{RAW_EPA}/2023_data_summary_spreadsheets/ghgp_data_2023.xlsx'
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)

# Sheet → subpart mapping
SHEETS = {
    'Onshore Oil & Gas Prod.': 'W-ONSH',
    'Gathering & Boosting': 'W-GB',
    'Transmission Pipelines': 'W-TRANS',
    'LDC - Direct Emissions': 'W-LDC',
}
og_facility_subparts = {}  # facility_id -> list of subparts it reports under

for sheet_name, subpart in SHEETS.items():
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[3])
    fid_idx = headers.index('Facility Id')
    for r in rows[4:]:
        if r[fid_idx] is None: break
        og_facility_subparts.setdefault(int(r[fid_idx]), []).append(subpart)

# W-PROC (Natural Gas Processing) lives in Direct Point Emitters
ws = wb['Direct Point Emitters']
rows = list(ws.iter_rows(values_only=True))
headers = list(rows[3])
fid_idx = headers.index('Facility Id')
sub_idx = headers.index('Industry Type (subparts)')
for r in rows[4:]:
    if r[fid_idx] is None: break
    if 'W-PROC' in str(r[sub_idx] or ''):
        og_facility_subparts.setdefault(int(r[fid_idx]), []).append('W-PROC')
wb.close()

og_ids = set(og_facility_subparts.keys())
print(f"Oil & Gas (Subpart W) facilities: {len(og_ids)}")

# ---------- 2. Join parent-company ownership ----------
with open_workbook(f'{RAW_EPA}/ghgp_data_parent_company.xlsb') as pwb:
    with pwb.get_sheet('2023') as sheet:
        rows = list(sheet.rows())
        headers = [c.v for c in rows[0]]
        data = [[c.v for c in r] for r in rows[1:]]
parent = pd.DataFrame(data, columns=headers)
parent['GHGRP FACILITY ID'] = parent['GHGRP FACILITY ID'].astype(int)

og_parent = parent[parent['GHGRP FACILITY ID'].isin(og_ids)].copy()

# Top parents in O&G (any-stake rule — same methodology we'll use GHGRP-wide)
og_by_parent = og_parent.groupby('PARENT COMPANY NAME')['GHGRP FACILITY ID'].nunique().sort_values(ascending=False)
top10_og = og_by_parent.head(10)
print("\nTop 10 O&G parents:")
print(top10_og.to_string())

top10_og_names = list(top10_og.index)
top10_og_facilities = set(og_parent[og_parent['PARENT COMPANY NAME'].isin(top10_og_names)]['GHGRP FACILITY ID'])

og_top10_share = 100 * len(top10_og_facilities) / len(og_ids)

# GHGRP-wide top 10 comparison (same method)
ghgrp_by_parent = parent.groupby('PARENT COMPANY NAME')['GHGRP FACILITY ID'].nunique().sort_values(ascending=False)
top10_ghgrp_names = list(ghgrp_by_parent.head(10).index)
top10_ghgrp_facilities = set(parent[parent['PARENT COMPANY NAME'].isin(top10_ghgrp_names)]['GHGRP FACILITY ID'])
ghgrp_top10_share = 100 * len(top10_ghgrp_facilities) / parent['GHGRP FACILITY ID'].nunique()

print(f"\nO&G top 10 share: {og_top10_share:.1f}%")
print(f"GHGRP-wide top 10 share: {ghgrp_top10_share:.1f}%")
print(f"Concentration multiplier: {og_top10_share/ghgrp_top10_share:.2f}x")

overlap = set(top10_og_names) & set(top10_ghgrp_names)
og_specific = set(top10_og_names) - set(top10_ghgrp_names)
print(f"\nShared top-10 names: {sorted(overlap)}")
print(f"O&G-specific top-10 (not in GHGRP-wide top 10): {sorted(og_specific)}")

# ---------- 3. Build facility-level ownership label for mapping ----------
# Each facility gets assigned to ONE parent for coloring — use the first of top 10 that owns it,
# else 'other' if no top-10 owner
parent_priority = {name: i for i, name in enumerate(top10_og_names)}

def top_owner(gid):
    owners = set(og_parent[og_parent['GHGRP FACILITY ID']==gid]['PARENT COMPANY NAME'])
    top10_owners = [n for n in owners if n in parent_priority]
    if not top10_owners:
        return 'other'
    # return the one with highest rank (smallest priority number)
    return sorted(top10_owners, key=lambda n: parent_priority[n])[0]

# ---------- 4. Load coordinates from existing processed file ----------
geo = pd.read_csv(f'{PROC}/ghgrp_2023_geo.csv')
geo['Facility Id'] = geo['Facility Id'].astype(int)
og_geo = geo[geo['Facility Id'].isin(og_ids)].copy()
og_geo['owner'] = og_geo['Facility Id'].apply(top_owner)
print(f"\nO&G facilities with coordinates: {len(og_geo)} of {len(og_ids)}")
print(f"Owner label counts:")
print(og_geo['owner'].value_counts())

# Save processed output
og_geo.to_csv(f'{PROC}/ghgrp_2023_og_facilities.csv', index=False)

# Save key numbers
with open(f'{PROC}/phase4_og_results.json', 'w') as f:
    json.dump({
        'og_total_facilities': int(len(og_ids)),
        'og_with_coords': int(len(og_geo)),
        'og_unique_parents': int(og_parent['PARENT COMPANY NAME'].nunique()),
        'og_top10_share_pct': float(og_top10_share),
        'ghgrp_wide_top10_share_pct': float(ghgrp_top10_share),
        'concentration_multiplier': float(og_top10_share / ghgrp_top10_share),
        'top10_og': {name: int(count) for name, count in top10_og.items()},
        'top10_ghgrp_wide': {name: int(ghgrp_by_parent[name]) for name in top10_ghgrp_names},
        'shared_names': sorted(overlap),
        'og_specific_names': sorted(og_specific),
    }, f, indent=2)
print(f"\nSaved: {PROC}/phase4_og_results.json")

# ---------- 5. Render map ----------
print("\nRendering...")
with open(f'{RAW_CB}/gz_2010_us_040_00_5m.json', encoding='latin-1') as f:
    sj = json.load(f)
srows = []
for feat in sj['features']:
    p = feat['properties']
    p['geometry'] = shape(feat['geometry'])
    srows.append(p)
states = gpd.GeoDataFrame(srows, crs='EPSG:4326')
conus_states = states[~states['NAME'].isin(['Alaska', 'Hawaii', 'Puerto Rico'])]

# CONUS extent filter on points
og_geo_conus = og_geo[
    (og_geo['Longitude'] >= -125) & (og_geo['Longitude'] <= -66) &
    (og_geo['Latitude'] >= 24) & (og_geo['Latitude'] <= 50)
].copy()

# Color palette for top 10
PALETTE = [
    '#e12828',  # ET - crimson
    '#2a6cd4',  # KM - blue
    '#11998e',  # Phillips 66 - teal
    '#e8822c',  # Targa - orange
    '#8f3fbd',  # Enterprise - purple
    '#c5a300',  # Enbridge - dark yellow
    '#9b1d20',  # ExxonMobil - dark red
    '#046b49',  # Williams - forest green
    '#d13d80',  # MPLX - magenta
    '#3e4d6b',  # EOG - slate
]
owner_colors = dict(zip(top10_og_names, PALETTE))
owner_colors['other'] = '#bfbfbf'

fig, ax = plt.subplots(figsize=(16, 9), dpi=150)
fig.patch.set_facecolor('#fafafa')
ax.set_facecolor('#fafafa')

# Basemap
conus_states.plot(ax=ax, color='#f0f0f0', edgecolor='#9a9a9a', linewidth=0.6, zorder=1)

# Plot others first (gray), then top 10 in reverse rank order (so #1 is on top)
for owner in ['other'] + list(reversed(top10_og_names)):
    subset = og_geo_conus[og_geo_conus['owner'] == owner]
    if len(subset) == 0: continue
    if owner == 'other':
        ax.scatter(subset['Longitude'], subset['Latitude'],
                   s=7, c=owner_colors[owner], alpha=0.5, linewidth=0, zorder=2)
    else:
        ax.scatter(subset['Longitude'], subset['Latitude'],
                   s=34, c=owner_colors[owner], alpha=0.9, zorder=3,
                   edgecolor='white', linewidth=0.5)

ax.set_xlim(-126, -65)
ax.set_ylim(23.5, 50)
ax.set_aspect(1.3)
ax.set_xticks([]); ax.set_yticks([])
for spine in ax.spines.values(): spine.set_visible(False)

# Title block
fig.text(0.06, 0.945, 'Oil & gas is more concentrated than GHGRP overall',
         fontsize=22, fontweight='bold', color='#111', family='sans-serif')
fig.text(0.06, 0.905,
         f'Top 10 parents control {og_top10_share:.1f}% of O&G facilities — {og_top10_share/ghgrp_top10_share:.1f}x the GHGRP-wide top 10 share ({ghgrp_top10_share:.1f}%).',
         fontsize=12.5, color='#444', family='sans-serif')

# Facility-count badges (metric cards)
def card(x, top, value, label, color='#111'):
    fig.text(x, top, value, fontsize=17, fontweight='bold', color=color)
    fig.text(x, top-0.022, label, fontsize=9.5, color='#444')

card(0.06, 0.855, f"{len(og_ids):,}", 'O&G facilities (Subpart W)', '#333')
card(0.22, 0.855, f"{og_parent['PARENT COMPANY NAME'].nunique():,}", 'unique O&G parents', '#333')
card(0.37, 0.855, f"{og_top10_share:.0f}%", 'top 10 share', '#c8292c')

# Legend: top 10 with counts
legend_elems = []
for owner in top10_og_names:
    count = int(top10_og[owner])
    # shorten names for legend
    display = owner.replace(' INC','').replace(' CORP','').replace(' LLC','').replace(' LP','').replace(' CO','').replace(' LTD','').title()
    legend_elems.append(
        Line2D([0],[0], marker='o', color='w', markerfacecolor=owner_colors[owner],
               markersize=9, label=f'{display}  ({count})')
    )
legend_elems.append(
    Line2D([0],[0], marker='o', color='w', markerfacecolor=owner_colors['other'],
           markersize=6, alpha=0.7, label=f'All other O&G operators')
)
# Legend placed BELOW the map axes so it never covers dots
leg = fig.legend(handles=legend_elems, loc='lower center', frameon=False,
                 fontsize=9.5, labelcolor='#222', bbox_to_anchor=(0.5, 0.17),
                 ncol=6, columnspacing=1.2, handletextpad=0.4)

# Footer (pushed lower to make room for legend)
fig.text(0.06, 0.075,
         'Scope: GHGRP Subpart W (petroleum and natural gas systems) — onshore production, gathering and boosting, processing, transmission, and LDCs.',
         fontsize=8.5, color='#666')
fig.text(0.06, 0.055,
         'Source: EPA GHGRP 2023 Data Summary Spreadsheets + Parent Company Dataset. Top 10 computed by unique-facility count, any-stake ownership rule.',
         fontsize=8.5, color='#666')
fig.text(0.06, 0.035,
         'Top 10 share = share of unique facilities in which any top-10 parent holds an ownership stake. Joint ventures count in both parents unless collapsed.',
         fontsize=8.5, color='#666')
fig.text(0.06, 0.015,
         'Analysis: @salamituns  |  Stratdevs (Tareony)',
         fontsize=9, color='#888', style='italic')

plt.subplots_adjust(left=0.03, right=0.97, top=0.80, bottom=0.22)

OUT = f'{VIZ}/ghgrp_oilgas_map_2023.png'
plt.savefig(OUT, dpi=200, bbox_inches='tight', facecolor='#fafafa')
print(f"Saved: {OUT}")
plt.close()
